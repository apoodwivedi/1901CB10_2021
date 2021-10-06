import genericpath
import os
import sys
import shutil
import csv
import openpyxl
from openpyxl import Workbook
os.system("cls")

input_file1 = "grades.csv"
input_file2 = "names-roll.csv"
input_file3 = "subjects_master.csv"
gradeValues = {
    "AA": 10,
    "AB": 9,
    "BB": 8,
    "BC": 7,
    "CC": 6,
    "CD": 5,
    "DD": 4,
    "F": 0,
    "I": 0
}
header = ["Sl No.", "Subject No.", "Subject Name",
          "L-T-P", "Credit", "Subject Type", "Grade"]
subjectName = os.path.join(os.getcwd(), input_file3)
studentName = os.path.join(os.getcwd(), input_file2)


def overallResult(rolln: str):
    spi = ["SPI"]
    cpi = ["CPI"]
    name = ["Name of Student"]
    semCredits = ["Semester wise Credit Taken"]
    totalCredits = ["Total Credits Taken"]
    semNumber = ["Semester No.", 1]
    numberOfSems = 0  # to find number of semesters of each student

    for i, line in enumerate(csv.reader(open(studentName))):
        if i > 0 and line[0] == rolln:
            name.append(line[1])  # add names

    for i, line in enumerate(csv.reader(open(input_file1))):
        if line[0] == rolln:
            # because not everyone has 8 semesters
            numberOfSems = max(numberOfSems, int(line[1]))

    numberOfSems += 1

    def stargradesFix(grade) -> str:  # DD*, FF* to DD,FF
        return grade.replace('*', '') if grade[len(grade) - 1] == '*' else grade

    for x in range(1, numberOfSems):
        creds = 0
        credsIntoGrades = 0

        for i, line in enumerate((csv.reader(open(input_file1)))):
            if i > 0:
                if line[0] == rolln:
                    if(int(line[1]) == x):
                        creds += int(line[3])
                        grade = stargradesFix(line[4].strip())
                        credsIntoGrades += int(line[3])*gradeValues[grade]

        if creds == 0:
            semCredits.append(0)
            spi.append(0)
        else:
            spi.append((credsIntoGrades/creds).__round__(2))
            semCredits.append(creds)

    cpivalues = spi[1]*semCredits[1]
    cumulativeCreds = semCredits[1]
    totalCredits.append(cumulativeCreds)

    cpi.append(spi[1])

    for s in range(2, numberOfSems):
        semNumber.append(s)
        cumulativeCreds += semCredits[s]
        totalCredits.append(cumulativeCreds)
        cpivalues += spi[s]*semCredits[s]
        cpi.append((cpivalues/cumulativeCreds).__round__(2))

    return name, semNumber, semCredits,  spi, totalCredits, cpi


def generate_marksheet():

    outputFolder = os.path.join(os.getcwd(), "output")
    if(os.path.exists(outputFolder)):
        shutil.rmtree(outputFolder)
    os.mkdir(outputFolder)

    for i, row in enumerate((csv.reader(open(input_file1)))):
        # data[2],[3] will be filled from input file 3
        data = [i, row[2], "", "", row[3], row[5], row[4].strip()]
        # file for each student i.e. rolln.xlsx
        resultFile = os.path.join(outputFolder, row[0] + ".xlsx")

        for j, jrow in enumerate((csv.reader(open(subjectName)))):
            if jrow[0] == data[1]:  # if line contains the same subject
                data[2] = jrow[1]    # store  ltp
                data[3] = jrow[2]    # store credits

        if i > 0:
            if not os.path.exists(resultFile):
                wrbk = openpyxl.Workbook()
                wrbk.save(resultFile)

            requiredSheet = "Sem " + str(row[1])  # creating semester sheets
            wb1 = openpyxl.load_workbook(resultFile)

            if not "Overall" in wb1.sheetnames:
                del wb1["Sheet"]
                overall = "Overall"
                wb1.create_sheet(overall)
                sheet = wb1[overall]
                name, semN, semCredits, spi, totalCredits, cpi = overallResult(
                    row[0])
                sheet.append(["Roll No.", row[0]])
                sheet.append(name)
                sheet.append(
                    ["Discipline", str(str(row[0])[4] + str(row[0])[5])])
                sheet.append(semN)
                sheet.append(semCredits)
                sheet.append(spi)
                sheet.append(totalCredits)
                sheet.append(cpi)
                wb1.save(resultFile)

            if not requiredSheet in wb1.sheetnames:
                wb1.create_sheet(requiredSheet)
                wb1[requiredSheet].append(header)
                wb1.save(resultFile)

            activeWb = wb1
            data[0] = activeWb[requiredSheet].max_row
            activeWb[requiredSheet].append(data)
            wb1.save(resultFile)

    return


generate_marksheet()
