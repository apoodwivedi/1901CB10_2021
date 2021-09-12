import os
import sys
import shutil
import csv
import openpyxl
from openpyxl import Workbook

os.system("cls")


input_file = "regtable_old.csv"


def output_by_subject():
    outputFolder = "output_by_subject"

    if(os.path.exists(outputFolder)):
        shutil.rmtree(outputFolder)
    os.mkdir(outputFolder)

    # reading csv file
    with open(input_file, 'r') as csvfile:
        csvreader = csv.reader(csvfile)
        fields = next(csvreader)
        outputFields = [fields[0], fields[1], fields[3], fields[8]]  # headers

        for row in csvreader:
            rows = [row[0], row[1], row[3], row[8]]

            output_file = os.path.join(outputFolder, row[3] + ".xlsx")

            if not os.path.exists(output_file):
                wb = Workbook()  # creating workbook to save excel files
                active = wb.active
                active.append(outputFields)
                wb.save(output_file)

            wb = openpyxl.load_workbook(output_file)
            act_wb = wb.active
            act_wb.append(rows)  # appending desired input
            wb.save(output_file)


def output_individual_roll():
    outputFolder = "output_individual_roll"

    # if output folder already exists remove it
    if(os.path.exists(outputFolder)):
        shutil.rmtree(outputFolder)
    os.mkdir(outputFolder)

    # reading csv file
    with open(input_file, 'r') as csvfile:
        csvreader = csv.reader(csvfile)
        fields = next(csvreader)
        outputFields = [fields[0], fields[1], fields[3], fields[8]]  # headers

        for row in csvreader:
            rows = [row[0], row[1], row[3], row[8]]

            output_file = os.path.join(outputFolder, row[0] + ".xlsx")

            if not os.path.exists(output_file):
                wb = Workbook()
                active = wb.active
                active.append(outputFields)
                wb.save(output_file)

            wb = openpyxl.load_workbook(output_file)
            act_wb = wb.active
            act_wb.append(rows)
            wb.save(output_file)


output_by_subject()
output_individual_roll()
